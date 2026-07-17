#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
agent_analyzer.py — автономный анализ директории upload:
  - .md файлы (бизнес-логика, цели, процессы)
  - Excel (.xlsx / .xls) с оптимизацией по памяти
  - естественные ключи и карта связей
  - итоговый отчёт PROJECT_ANALYSIS_REPORT.md
"""

from __future__ import annotations

import os
import re
import sys
import traceback
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

# ---------------------------------------------------------------------------
# Константы
# ---------------------------------------------------------------------------
# Скрипт живёт в корне проекта; исходники — только в upload/
PROJECT_ROOT = Path(__file__).resolve().parent
ROOT = PROJECT_ROOT / "upload"
REPORT_PATH = PROJECT_ROOT / "tool-results" / "PROJECT_ANALYSIS_REPORT.md"
ERRORS_LOG = PROJECT_ROOT / "db" / "analysis_errors.log"
HEADER_SCAN_ROWS = 10
SAMPLE_DATA_ROWS = 100

# гарантируем каталоги артефактов
REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
ERRORS_LOG.parent.mkdir(parents=True, exist_ok=True)

# Ключевые слова для извлечения бизнес-логики из .md
MD_KEYWORDS = [
    "цель", "задач", "автоматизац", "отчет", "отчёт", "интеграц", "etl",
    "качество данных", "процесс", "проблем", "решени", "архитектур",
    "хранилищ", "импорт", "экспорт", "дашборд", "kpi", "сотрудник",
    "табель", "площадк", "биле", "миграция", "pipeline", "масштабир",
    "валидац", "справочник", "реестр", "учет", "учёт", "мониторинг",
]

# Паттерны естественных ключей (очищенные имена колонок)
NATURAL_KEY_PATTERNS = [
    (r"табельн", "табельный_номер"),
    (r"tab(el)?n", "табельный_номер"),
    (r"\bфио\b|ф_и_о|фамилия_имя|fio", "ФИО"),
    (r"паспорт|passport", "паспорт"),
    (r"\bинн\b|\binn\b", "ИНН"),
    (r"снилс|snils", "СНИЛС"),
    (r"^id$|_id$|^id_|идентификатор|unique_id", "ID"),
    (r"артикул|sku|article", "артикул"),
    (r"^код$|код_|_код|code$|_code", "код"),
    (r"номер_договор|contract", "номер_договора"),
    (r"дата_рожден|др_|birth", "дата_рождения"),
    (r"guid|uuid", "GUID"),
    (r"таб_номер|табномер|pers_number|personnel", "табельный_номер"),
]


def safe_print(text: str, *, file=None) -> None:
    """Печать с безопасной кодировкой для Windows-консоли (cp1251 и т.п.)."""
    stream = file or sys.stdout
    try:
        stream.write(text + "\n")
        stream.flush()
    except UnicodeEncodeError:
        enc = getattr(stream, "encoding", None) or "utf-8"
        stream.buffer.write((text + "\n").encode(enc, errors="replace"))
        stream.flush()


def log_error(message: str, file_name: str = "", sheet: str = "") -> None:
    """Дописать ошибку в errors.log с меткой времени."""
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    parts = [f"[{ts}]"]
    if file_name:
        parts.append(f"file={file_name}")
    if sheet:
        parts.append(f"sheet={sheet}")
    parts.append(message)
    line = " | ".join(parts)
    try:
        with open(ERRORS_LOG, "a", encoding="utf-8") as f:
            f.write(line + "\n")
    except OSError:
        safe_print(f"WARNING: cannot write errors.log: {line}", file=sys.stderr)
    safe_print(f"ERROR: {line}", file=sys.stderr)


def clean_column_name(name: Any) -> str:
    """Очистка названия колонки по правилам ТЗ."""
    if name is None:
        return ""
    s = str(name).strip().lower()
    s = re.sub(r"[^\w]+", "_", s, flags=re.UNICODE)
    s = re.sub(r"_+", "_", s)
    s = s.strip("_")
    return s


def normalize_for_match(name: str) -> str:
    """Нормализация имени колонки для поиска пересечений."""
    return clean_column_name(name).replace("_", "")


def read_text_file(path: Path) -> str:
    """Читать текстовый файл UTF-8 с fallback на cp1251."""
    for enc in ("utf-8", "utf-8-sig", "cp1251"):
        try:
            return path.read_text(encoding=enc)
        except UnicodeDecodeError:
            continue
        except OSError as e:
            log_error(f"read failed: {e}", str(path.relative_to(ROOT)))
            return ""
    log_error("encoding failed for all codecs", str(path.relative_to(ROOT)))
    return ""


# ===========================================================================
# 2.1 Анализ .md
# ===========================================================================

def extract_md_insights(text: str, rel_path: str) -> dict[str, list[str]]:
    """Извлечь релевантные фрагменты по ключевым словам."""
    insights: dict[str, list[str]] = {
        "goals": [],
        "tasks": [],
        "processes": [],
        "problems": [],
        "solutions": [],
        "other": [],
    }
    if not text.strip():
        return insights

    lines = text.splitlines()
    for i, line in enumerate(lines):
        low = line.lower().strip()
        if not low or len(low) < 15:
            continue
        matched = False
        for kw in MD_KEYWORDS:
            if kw in low:
                snippet = line.strip()
                if len(snippet) > 300:
                    snippet = snippet[:297] + "..."
                if any(x in low for x in ("цель", "goal", "purpose")):
                    insights["goals"].append(f"[{rel_path}] {snippet}")
                elif any(x in low for x in ("задач", "task")):
                    insights["tasks"].append(f"[{rel_path}] {snippet}")
                elif any(x in low for x in ("проблем", "ошибк", "issue", "bug", "долг")):
                    insights["problems"].append(f"[{rel_path}] {snippet}")
                elif any(x in low for x in ("решени", "fix", "исправ", "предложен")):
                    insights["solutions"].append(f"[{rel_path}] {snippet}")
                elif any(x in low for x in ("процесс", "etl", "интеграц", "pipeline", "импорт")):
                    insights["processes"].append(f"[{rel_path}] {snippet}")
                else:
                    insights["other"].append(f"[{rel_path}] {snippet}")
                matched = True
                break
        if matched:
            continue
        # Заголовки с бизнес-смыслом
        if line.startswith("#") and any(
            k in low for k in ("архитектур", "описани", "обзор", "summary", "overview")
        ):
            insights["other"].append(f"[{rel_path}] {line.strip()[:300]}")

    # Дедупликация с сохранением порядка
    for key in insights:
        seen: set[str] = set()
        uniq: list[str] = []
        for item in insights[key]:
            if item not in seen:
                seen.add(item)
                uniq.append(item)
        insights[key] = uniq[:40]  # лимит на категорию
    return insights


def analyze_markdown_files() -> tuple[dict[str, list[str]], list[str]]:
    """Рекурсивный обход .md и агрегация инсайтов."""
    aggregated: dict[str, list[str]] = {
        "goals": [],
        "tasks": [],
        "processes": [],
        "problems": [],
        "solutions": [],
        "other": [],
    }
    md_files: list[str] = []

    for path in sorted(ROOT.rglob("*.md")):
        # Пропускаем собственные артефакты
        if path.name in ("PROJECT_ANALYSIS_REPORT.md",):
            continue
        if path.name.startswith("~$"):
            continue
        rel = str(path.relative_to(ROOT)).replace("\\", "/")
        md_files.append(rel)
        try:
            text = read_text_file(path)
            insights = extract_md_insights(text, rel)
            for key in aggregated:
                aggregated[key].extend(insights[key])
        except Exception as e:
            log_error(f"md analysis: {e}\n{traceback.format_exc()}", rel)

    # глобальная дедупликация
    for key in aggregated:
        seen: set[str] = set()
        uniq: list[str] = []
        for item in aggregated[key]:
            if item not in seen:
                seen.add(item)
                uniq.append(item)
        aggregated[key] = uniq

    return aggregated, md_files


def build_improved_project_idea(insights: dict[str, list[str]], md_count: int) -> str:
    """Сгенерировать раздел «Улучшенная идея проекта»."""
    lines: list[str] = []
    lines.append("## 1. Улучшенная идея проекта")
    lines.append("")
    lines.append(
        f"На основе анализа **{md_count}** markdown-документов сформирована "
        "улучшенная концепция единого аналитического контура кадрово-производственного учёта "
        "(OMiK / VSM / ВелесстройМонтаж)."
    )
    lines.append("")
    lines.append("### 1.1. Суть проекта")
    lines.append("")
    lines.append(
        "Построить **единое хранилище данных (Data Warehouse / Lakehouse)** для HR-, "
        "производственных и логистических данных компании: ежедневный учёт персонала на площадках, "
        "календари прилёт/вылет, реестры билетов, приём/перевод/увольнение, оценки, "
        "справочники должностей и площадок. Источники — разнородные Excel-файлы и выгрузки 1С; "
        "потребители — дашборды KPI, drill-down по сотрудникам и площадкам, отчёты для руководства."
    )
    lines.append("")
    lines.append("### 1.2. Ключевые цели (из документации)")
    lines.append("")
    if insights["goals"]:
        for g in insights["goals"][:15]:
            lines.append(f"- {g}")
    else:
        lines.append("- Автоматизация сбора и консолидации кадрово-производственных данных.")
        lines.append("- Единый профиль сотрудника и прозрачная аналитика по площадкам.")
        lines.append("- Контроль качества данных и снижение ручного труда в Excel/VBA.")
    lines.append("")
    lines.append("### 1.3. Бизнес-процессы и задачи")
    lines.append("")
    if insights["tasks"] or insights["processes"]:
        for item in (insights["tasks"] + insights["processes"])[:20]:
            lines.append(f"- {item}")
    else:
        lines.append("- ETL-загрузка Excel → staging → нормализация → витрины.")
        lines.append("- Сопоставление сотрудников между источниками по естественному ключу.")
        lines.append("- Построение KPI и drill-down отчётов.")
    lines.append("")
    lines.append("### 1.4. Выявленные проблемы")
    lines.append("")
    if insights["problems"]:
        for p in insights["problems"][:15]:
            lines.append(f"- {p}")
    else:
        lines.append("- Разрозненные Excel-файлы без единого ключа сотрудника.")
        lines.append("- Ручные макросы VBA, слабый контроль качества данных.")
        lines.append("- Дублирование справочников и расхождения между площадками.")
    lines.append("")
    lines.append("### 1.5. Предложенные решения из документации")
    lines.append("")
    if insights["solutions"]:
        for s in insights["solutions"][:15]:
            lines.append(f"- {s}")
    else:
        lines.append("- Централизованный backend + фронтенд-дашборды.")
        lines.append("- Потоковая/пакетная загрузка с валидацией.")
    lines.append("")
    lines.append("### 1.6. Улучшенная целевая архитектура (best practices)")
    lines.append("")
    lines.append(
        "Объединяя лучшие практики из описаний проектов OMiK/VSM и современные подходы "
        "к инженерии данных, рекомендуется следующий контур:"
    )
    lines.append("")
    lines.append("1. **Ingestion (ETL/ELT)**")
    lines.append("   - Пакетная загрузка Excel/1С в staging (parquet/БД) с версионированием файлов.")
    lines.append("   - Идемпотентные job'ы, checksum файлов, журнал загрузок (audit log).")
    lines.append("   - При появлении потоковых источников (API, SSE) — CDC / streaming layer.")
    lines.append("")
    lines.append("2. **Качество данных (DQ) на всех этапах**")
    lines.append("   - Schema validation, обязательные поля, форматы ИНН/СНИЛС/дат.")
    lines.append("   - Правила уникальности по `employee_uid` (см. раздел 3).")
    lines.append("   - Quarantine для «грязных» строк + алерты.")
    lines.append("")
    lines.append("3. **Модель данных**")
    lines.append("   - Справочники: площадки, должности, подразделения, классификации.")
    lines.append("   - Факты: ежедневный учёт, перемещения, билеты, оценки.")
    lines.append("   - SCD2 для медленно меняющихся измерений (должность, площадка).")
    lines.append("")
    lines.append("4. **Масштабируемость**")
    lines.append("   - Разделение raw / curated / marts.")
    lines.append("   - Инкрементальная обработка по дате файла / дате учёта.")
    lines.append("   - Горизонтальное масштабирование воркеров ETL при росте объёма.")
    lines.append("")
    lines.append("5. **Потребление**")
    lines.append("   - API + дашборды KPI, drill-down, экспорт.")
    lines.append("   - Ролевая модель доступа (площадка / подразделение).")
    lines.append("")
    if insights["other"]:
        lines.append("### 1.7. Дополнительные сигналы из документации")
        lines.append("")
        for o in insights["other"][:12]:
            lines.append(f"- {o}")
        lines.append("")
    return "\n".join(lines)


# ===========================================================================
# 2.2 Анализ Excel
# ===========================================================================

def score_header_row(values: list[Any]) -> float:
    """Эвристика: меньше пустых, больше строковых (нечисловых) значений."""
    if not values:
        return -1.0
    non_empty = 0
    string_like = 0
    for v in values:
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        non_empty += 1
        if isinstance(v, str):
            s = v.strip()
            # числа в виде строк не считаем «строковыми заголовками»
            try:
                float(s.replace(",", ".").replace(" ", ""))
                continue
            except ValueError:
                string_like += 1
        elif isinstance(v, (int, float)):
            pass
        else:
            string_like += 1
    if non_empty == 0:
        return -1.0
    empty_penalty = (len(values) - non_empty) / max(len(values), 1)
    return string_like * 2.0 + non_empty * 1.0 - empty_penalty * 5.0


def analyze_xlsx_sheet(wb_path: Path, sheet_name: str, ws: Any) -> dict[str, Any] | None:
    """Анализ одного листа openpyxl (read_only)."""
    try:
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        if max_row == 0 or max_col == 0:
            return {
                "file": str(wb_path.relative_to(ROOT)).replace("\\", "/"),
                "sheet": sheet_name,
                "header_row": None,
                "total_rows": 0,
                "columns": [],
                "note": "пустой лист",
            }

        # Читаем первые HEADER_SCAN_ROWS строк для поиска заголовка
        rows_sample: list[tuple[int, list[Any]]] = []
        for i, row in enumerate(ws.iter_rows(max_row=HEADER_SCAN_ROWS, values_only=True), start=1):
            rows_sample.append((i, list(row) if row else []))

        best_idx = None
        best_score = -1.0
        best_values: list[Any] = []
        for idx, vals in rows_sample:
            sc = score_header_row(vals)
            if sc > best_score:
                best_score = sc
                best_idx = idx
                best_values = vals

        columns: list[str] = []
        if best_idx is not None and best_score > 0:
            raw_names = [v if v is not None else f"col_{i+1}" for i, v in enumerate(best_values)]
            # обрезаем хвостовые полностью пустые
            while raw_names and (raw_names[-1] is None or str(raw_names[-1]).strip() == "" or
                                 (isinstance(raw_names[-1], str) and raw_names[-1].startswith("col_") and
                                  best_values[len(raw_names)-1] is None)):
                # более простая очистка хвоста
                if best_values[len(raw_names) - 1] is None or (
                    isinstance(best_values[len(raw_names) - 1], str)
                    and not str(best_values[len(raw_names) - 1]).strip()
                ):
                    raw_names.pop()
                    continue
                break
            seen_names: dict[str, int] = {}
            for i, raw in enumerate(raw_names):
                if best_values[i] is None or (isinstance(best_values[i], str) and not str(best_values[i]).strip()):
                    cname = f"unnamed_{i+1}"
                else:
                    cname = clean_column_name(raw) or f"unnamed_{i+1}"
                if cname in seen_names:
                    seen_names[cname] += 1
                    cname = f"{cname}_{seen_names[cname]}"
                else:
                    seen_names[cname] = 1
                columns.append(cname)

        return {
            "file": str(wb_path.relative_to(ROOT)).replace("\\", "/"),
            "sheet": sheet_name,
            "header_row": best_idx,
            "total_rows": max_row,
            "columns": columns,
            "note": "",
        }
    except Exception as e:
        log_error(f"xlsx sheet: {e}\n{traceback.format_exc()}", str(wb_path), sheet_name)
        return None


def analyze_xlsx(path: Path) -> list[dict[str, Any]]:
    """Анализ .xlsx через openpyxl read_only."""
    results: list[dict[str, Any]] = []
    rel = str(path.relative_to(ROOT)).replace("\\", "/")
    try:
        import openpyxl
    except ImportError as e:
        log_error(f"openpyxl import failed: {e}", rel)
        return results

    wb = None
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
                info = analyze_xlsx_sheet(path, sheet_name, ws)
                if info:
                    results.append(info)
            except Exception as e:
                log_error(f"sheet error: {e}\n{traceback.format_exc()}", rel, sheet_name)
    except Exception as e:
        log_error(f"xlsx open failed: {e}\n{traceback.format_exc()}", rel)
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
    return results


def analyze_xls(path: Path) -> list[dict[str, Any]]:
    """Анализ .xls через xlrd (read-only)."""
    results: list[dict[str, Any]] = []
    rel = str(path.relative_to(ROOT)).replace("\\", "/")
    try:
        import xlrd
    except ImportError as e:
        log_error(f"xlrd import failed: {e}", rel)
        return results

    try:
        book = xlrd.open_workbook(str(path), on_demand=True)
    except Exception as e:
        log_error(f"xls open failed: {e}\n{traceback.format_exc()}", rel)
        return results

    try:
        for sheet_name in book.sheet_names():
            try:
                sh = book.sheet_by_name(sheet_name)
                max_row = sh.nrows
                if max_row == 0:
                    results.append({
                        "file": rel,
                        "sheet": sheet_name,
                        "header_row": None,
                        "total_rows": 0,
                        "columns": [],
                        "note": "пустой лист",
                    })
                    continue

                scan_n = min(HEADER_SCAN_ROWS, max_row)
                best_idx = None
                best_score = -1.0
                best_values: list[Any] = []
                for r in range(scan_n):
                    vals = [sh.cell_value(r, c) for c in range(sh.ncols)]
                    sc = score_header_row(vals)
                    if sc > best_score:
                        best_score = sc
                        best_idx = r + 1  # 1-based
                        best_values = vals

                columns: list[str] = []
                if best_idx is not None and best_score > 0:
                    seen_names: dict[str, int] = {}
                    # обрезаем хвостовые пустые
                    end = len(best_values)
                    while end > 0:
                        v = best_values[end - 1]
                        if v is None or v == "" or (isinstance(v, str) and not v.strip()):
                            end -= 1
                        else:
                            break
                    for i, raw in enumerate(best_values[:end]):
                        if raw is None or raw == "" or (isinstance(raw, str) and not str(raw).strip()):
                            cname = f"unnamed_{i+1}"
                        else:
                            cname = clean_column_name(raw) or f"unnamed_{i+1}"
                        if cname in seen_names:
                            seen_names[cname] += 1
                            cname = f"{cname}_{seen_names[cname]}"
                        else:
                            seen_names[cname] = 1
                        columns.append(cname)

                results.append({
                    "file": rel,
                    "sheet": sheet_name,
                    "header_row": best_idx,
                    "total_rows": max_row,
                    "columns": columns,
                    "note": "",
                })
            except Exception as e:
                log_error(f"xls sheet: {e}\n{traceback.format_exc()}", rel, sheet_name)
            finally:
                try:
                    book.unload_sheet(sheet_name)
                except Exception:
                    pass
    finally:
        try:
            book.release_resources()
        except Exception:
            pass

    return results


def find_excel_files() -> list[Path]:
    """Найти .xlsx и .xls, пропуская временные ~$ файлы."""
    files: list[Path] = []
    for pattern in ("*.xlsx", "*.xls"):
        for p in ROOT.rglob(pattern):
            if p.name.startswith("~$"):
                continue
            if p.suffix.lower() == ".xls" and p.suffix.lower() != ".xlsx":
                # rglob *.xls может не захватить xlsx, но на всякий случай
                pass
            # не берём .xlsm/.xlsb через этот фильтр — только явные расширения
            if p.suffix.lower() not in (".xlsx", ".xls"):
                continue
            files.append(p)
    # уникальные, отсортированные
    uniq = sorted(set(files), key=lambda x: str(x).lower())
    return uniq


def analyze_all_excel() -> list[dict[str, Any]]:
    """Обойти все Excel и собрать метаданные листов."""
    all_sheets: list[dict[str, Any]] = []
    files = find_excel_files()
    safe_print(f"Found {len(files)} Excel files (.xlsx/.xls)")
    for i, path in enumerate(files, 1):
        rel = str(path.relative_to(ROOT)).replace("\\", "/")
        try:
            safe_print(f"[{i}/{len(files)}] {rel}")
        except Exception:
            safe_print(f"[{i}/{len(files)}] <unprintable path>")
        try:
            suffix = path.suffix.lower()
            if suffix == ".xlsx":
                sheets = analyze_xlsx(path)
            elif suffix == ".xls":
                sheets = analyze_xls(path)
            else:
                continue
            all_sheets.extend(sheets)
        except Exception as e:
            log_error(f"file-level: {e}\n{traceback.format_exc()}", rel)
    return all_sheets


# ===========================================================================
# 2.3 Естественные ключи и связи
# ===========================================================================

def detect_natural_keys(sheets: list[dict[str, Any]]) -> dict[str, Any]:
    """Найти колонки-кандидаты в естественные ключи."""
    found: dict[str, list[dict[str, str]]] = defaultdict(list)
    # key_type -> list of {file, sheet, column}

    for sh in sheets:
        for col in sh.get("columns") or []:
            col_l = col.lower()
            for pattern, key_type in NATURAL_KEY_PATTERNS:
                if re.search(pattern, col_l, re.IGNORECASE):
                    found[key_type].append({
                        "file": sh["file"],
                        "sheet": sh["sheet"],
                        "column": col,
                    })
                    break

    return dict(found)


def propose_composite_key(natural_keys: dict[str, Any], sheets: list[dict[str, Any]]) -> str:
    """Предложить составной уникальный ID."""
    has_fio = "ФИО" in natural_keys
    has_tab = "табельный_номер" in natural_keys
    has_inn = "ИНН" in natural_keys
    has_birth = "дата_рождения" in natural_keys
    has_passport = "паспорт" in natural_keys
    has_snils = "СНИЛС" in natural_keys

    # Частота появления колонок по нормализованным именам
    col_freq: dict[str, int] = defaultdict(int)
    for sh in sheets:
        for c in sh.get("columns") or []:
            col_freq[normalize_for_match(c)] += 1

    lines: list[str] = []
    lines.append("### Предложенный составной ключ: `employee_uid`")
    lines.append("")

    if has_inn and has_tab:
        formula = "`SHA256(ИНН + '|' + табельный_номер)`"
        reason = (
            "ИНН однозначно идентифицирует физлицо в РФ, табельный номер — "
            "учётную запись в кадровой системе; вместе они устойчивы к опечаткам в ФИО."
        )
    elif has_snils:
        formula = "`SHA256(СНИЛС)` (при наличии) или `SHA256(ИНН)`"
        reason = "СНИЛС/ИНН — государственные уникальные идентификаторы."
    elif has_fio and has_birth:
        formula = "`SHA256(normalize(ФИО) + '|' + Дата_рождения + '|' + coalesce(табельный_номер,''))`"
        reason = (
            "ФИО + дата рождения — классический естественный ключ при отсутствии ИНН; "
            "добавление табельного номера снижает коллизии однофамильцев-тёзок."
        )
    elif has_fio and has_tab:
        formula = "`SHA256(normalize(ФИО) + '|' + табельный_номер)`"
        reason = (
            "ФИО и табельный номер встречаются в большинстве источников "
            "(ежедневный учёт, календари, билеты, приём/увольнение)."
        )
    elif has_passport:
        formula = "`SHA256(серия_номер_паспорта)`"
        reason = "Паспортные данные уникальны, но чувствительны — хранить только хеш."
    else:
        formula = (
            "`SHA256(normalize(ФИО) + '|' + normalize(площадка|подразделение) + '|' + дата_события)`"
        )
        reason = (
            "При слабой идентификационной модели — временный суррогатный ключ "
            "с последующим ручным matching / golden record."
        )

    lines.append(f"**Формула:** {formula}")
    lines.append("")
    lines.append(f"**Обоснование надёжности:** {reason}")
    lines.append("")
    lines.append("**Правила нормализации ФИО перед хешированием:**")
    lines.append("- верхний регистр, ё→е, удаление лишних пробелов и инициалов-точек;")
    lines.append("- единый порядок: Фамилия Имя Отчество;")
    lines.append("- хранение `employee_uid` во всех витринах как FK к измерению `dim_employee`.")
    lines.append("")
    lines.append(
        f"**Наблюдаемость ключей в данных:** "
        f"ФИО={'да' if has_fio else 'нет'}, "
        f"табельный={'да' if has_tab else 'нет'}, "
        f"ИНН={'да' if has_inn else 'нет'}, "
        f"СНИЛС={'да' if has_snils else 'нет'}, "
        f"паспорт={'да' if has_passport else 'нет'}, "
        f"дата_рождения={'да' if has_birth else 'нет'}."
    )
    lines.append("")
    return "\n".join(lines)


def build_relationship_map(sheets: list[dict[str, Any]]) -> list[dict[str, str]]:
    """Карта связей по пересечению имён колонок между листами разных файлов."""
    # normalized_col -> list of (file, sheet, original_col)
    index: dict[str, list[tuple[str, str, str]]] = defaultdict(list)
    for sh in sheets:
        file_s = sh["file"]
        sheet_s = sh["sheet"]
        seen_local: set[tuple[str, str, str]] = set()
        for col in sh.get("columns") or []:
            norm = normalize_for_match(col)
            if not norm or norm.startswith("unnamed") or len(norm) < 2:
                continue
            # пропускаем слишком общие имена
            if norm in ("дата", "date", "номер", "name", "имя", "примечание", "комментарий",
                        "сумма", "колво", "количество", "статус", "тип", "год", "месяц"):
                continue
            key = (file_s, sheet_s, norm)
            if key in seen_local:
                continue
            seen_local.add(key)
            index[norm].append((file_s, sheet_s, col))

    relations: list[dict[str, str]] = []
    # Для каждого нормализованного имени — связи между разными файлами
    for norm, locations in sorted(index.items(), key=lambda x: -len(x[1])):
        if len(locations) < 2:
            continue
        # уникальные файлы
        files = {loc[0] for loc in locations}
        if len(files) < 2:
            # связи внутри одного файла между листами тоже полезны
            sheets_set = {(loc[0], loc[1]) for loc in locations}
            if len(sheets_set) < 2:
                continue

        # берём пары: первый как «источник справочника», остальные как ссылающиеся
        # приоритет: папки Справочники / Базы как target (измерение)
        def priority(loc: tuple[str, str, str]) -> int:
            f = loc[0].lower()
            if "справочник" in f or "баз" in f:
                return 0
            if "ежедневн" in f:
                return 2
            return 1

        locs_sorted = sorted(locations, key=priority)
        # ограничиваем число рёбер
        hub = locs_sorted[0]
        for other in locs_sorted[1:]:
            if hub[0] == other[0] and hub[1] == other[1]:
                continue
            # тип связи: ключи-кандидаты → скорее 1:N или 1:1
            rel_type = "пересечение имён (потенциальный FK)"
            col_l = hub[2].lower()
            if re.search(r"табельн|инн|снилс|паспорт|^id$|guid|фио", col_l):
                rel_type = "потенциальный ключ / FK (1:N или 1:1)"
            relations.append({
                "src_file": other[0],
                "src_sheet": other[1],
                "src_col": other[2],
                "tgt_file": hub[0],
                "tgt_sheet": hub[1],
                "tgt_col": hub[2],
                "type": rel_type,
                "norm": norm,
            })

    # ограничим отчёт топ-N наиболее полезных
    # приоритет: ключи, затем частота
    def rel_score(r: dict[str, str]) -> int:
        score = 0
        if "ключ" in r["type"] or "FK" in r["type"]:
            score += 10
        for token in ("табель", "фио", "инн", "снилс", "площад", "должност", "подраздел"):
            if token in r["norm"]:
                score += 5
        return score

    relations.sort(key=lambda r: (-rel_score(r), r["norm"], r["src_file"]))
    # дедуп по (src, tgt, norm)
    seen: set[tuple] = set()
    uniq: list[dict[str, str]] = []
    for r in relations:
        k = (r["src_file"], r["src_sheet"], r["tgt_file"], r["tgt_sheet"], r["norm"])
        if k in seen:
            continue
        seen.add(k)
        uniq.append(r)
    return uniq[:200]


# ===========================================================================
# 2.4 Генерация отчёта
# ===========================================================================

def md_escape_cell(text: str) -> str:
    """Экранирование для ячеек markdown-таблицы."""
    return str(text).replace("|", "\\|").replace("\n", " ")


def generate_report(
    insights: dict[str, list[str]],
    md_files: list[str],
    sheets: list[dict[str, Any]],
    natural_keys: dict[str, Any],
    composite_key_md: str,
    relations: list[dict[str, str]],
) -> None:
    """Записать PROJECT_ANALYSIS_REPORT.md."""
    excel_files = sorted({s["file"] for s in sheets})
    lines: list[str] = []
    lines.append("# PROJECT ANALYSIS REPORT")
    lines.append("")
    lines.append(f"**Дата генерации:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"**Корневая директория:** `{ROOT}`")
    lines.append("")
    lines.append("### Сводка")
    lines.append("")
    lines.append(f"| Метрика | Значение |")
    lines.append(f"|---|---|")
    lines.append(f"| Markdown-файлов | {len(md_files)} |")
    lines.append(f"| Excel-файлов (.xlsx/.xls) | {len(excel_files)} |")
    lines.append(f"| Листов проанализировано | {len(sheets)} |")
    lines.append(f"| Связей (пересечений колонок) | {len(relations)} |")
    lines.append("")
    lines.append("---")
    lines.append("")

    # 1. Улучшенная идея
    lines.append(build_improved_project_idea(insights, len(md_files)))
    lines.append("---")
    lines.append("")

    # 2. Словарь данных
    lines.append("## 2. Словарь данных")
    lines.append("")
    lines.append(
        "Для каждого листа Excel: путь относительно корня, лист, строка заголовка (1-based), "
        "общее число строк, очищенные имена колонок."
    )
    lines.append("")

    # Группируем по файлу для читаемости
    by_file: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for s in sheets:
        by_file[s["file"]].append(s)

    for file_path in sorted(by_file.keys()):
        lines.append(f"### Файл: `{file_path}`")
        lines.append("")
        lines.append("| Лист | Строка заголовка | Всего строк | Колонки |")
        lines.append("|---|---:|---:|---|")
        for s in by_file[file_path]:
            cols = s.get("columns") or []
            if len(cols) > 40:
                cols_str = ", ".join(cols[:40]) + f", … (+{len(cols)-40})"
            else:
                cols_str = ", ".join(cols) if cols else "—"
            hr = s.get("header_row") if s.get("header_row") is not None else "—"
            note = f" _{s['note']}_" if s.get("note") else ""
            lines.append(
                f"| {md_escape_cell(s['sheet'])} | {hr} | {s.get('total_rows', 0)} | "
                f"{md_escape_cell(cols_str)}{note} |"
            )
        lines.append("")

    # Полная таблица-сводка (компактная)
    lines.append("### Сводная таблица (все листы)")
    lines.append("")
    lines.append("| Файл | Лист | Строка заголовка | Всего строк | Число колонок |")
    lines.append("|---|---|---:|---:|---:|")
    for s in sheets:
        hr = s.get("header_row") if s.get("header_row") is not None else "—"
        ncol = len(s.get("columns") or [])
        lines.append(
            f"| {md_escape_cell(s['file'])} | {md_escape_cell(s['sheet'])} | "
            f"{hr} | {s.get('total_rows', 0)} | {ncol} |"
        )
    lines.append("")
    lines.append("---")
    lines.append("")

    # 3. Уникальные идентификаторы
    lines.append("## 3. Уникальные идентификаторы")
    lines.append("")
    lines.append("### Найденные естественные ключи")
    lines.append("")
    if not natural_keys:
        lines.append("_Явные естественные ключи по названиям колонок не обнаружены._")
        lines.append("")
    else:
        for key_type, locs in sorted(natural_keys.items(), key=lambda x: x[0]):
            lines.append(f"#### {key_type} ({len(locs)} вхождений)")
            lines.append("")
            lines.append("| Файл | Лист | Колонка |")
            lines.append("|---|---|---|")
            # ограничим вывод
            for loc in locs[:50]:
                lines.append(
                    f"| {md_escape_cell(loc['file'])} | {md_escape_cell(loc['sheet'])} | "
                    f"{md_escape_cell(loc['column'])} |"
                )
            if len(locs) > 50:
                lines.append(f"| … | … | _ещё {len(locs)-50}_ |")
            lines.append("")

    lines.append(composite_key_md)
    lines.append("---")
    lines.append("")

    # 4. Карта связей
    lines.append("## 4. Карта связей между файлами")
    lines.append("")
    lines.append(
        "Связи построены по пересечению нормализованных имён колонок "
        "(без учёта регистра и разделителей). "
        "Справочники и базы приоритетно выступают как целевые (измерения)."
    )
    lines.append("")

    if not relations:
        lines.append("_Пересечений колонок между разными файлами/листами не найдено._")
        lines.append("")
    else:
        lines.append("| Исходный файл | Лист | Колонка | → | Целевой файл | Лист | Колонка | Тип связи |")
        lines.append("|---|---|---|---|---|---|---|---|")
        for r in relations:
            lines.append(
                f"| {md_escape_cell(r['src_file'])} | {md_escape_cell(r['src_sheet'])} | "
                f"{md_escape_cell(r['src_col'])} | → | "
                f"{md_escape_cell(r['tgt_file'])} | {md_escape_cell(r['tgt_sheet'])} | "
                f"{md_escape_cell(r['tgt_col'])} | {md_escape_cell(r['type'])} |"
            )
        lines.append("")

        # Текстовая диаграмма хабов
        lines.append("### Текстовая диаграмма ключевых хабов")
        lines.append("")
        lines.append("```")
        lines.append("dim_employee (employee_uid)")
        lines.append("    ^")
        lines.append("    |-- Ежедневный учёт (ФИО / табельный)")
        lines.append("    |-- Календарь прилёт-вылет")
        lines.append("    |-- Приём / Перевод / Увольнение")
        lines.append("    |-- Реестры билетов")
        lines.append("    |-- Оценки / Трудоустройство")
        lines.append("")
        lines.append("dim_worksite / dim_department")
        lines.append("    ^")
        lines.append("    |-- Справочники/Площадки_Регион")
        lines.append("    |-- Справочники/Подразделение_Участки*")
        lines.append("    |-- факты ежедневного учёта (площадка, участок)")
        lines.append("")
        lines.append("dim_position")
        lines.append("    ^")
        lines.append("    |-- Классификация_по_должности / Фактические_должности_ОК")
        lines.append("    |-- факты (должность)")
        lines.append("```")
        lines.append("")

    # Приложение: список md
    lines.append("---")
    lines.append("")
    lines.append("## Приложение A. Проанализированные Markdown-файлы")
    lines.append("")
    for m in md_files:
        lines.append(f"- `{m}`")
    lines.append("")
    lines.append("## Приложение B. Примечания по методологии")
    lines.append("")
    lines.append("- Excel читался в режиме `read_only` (openpyxl) / `on_demand` (xlrd), без полной загрузки в pandas.")
    lines.append(f"- Строка заголовка выбиралась среди первых {HEADER_SCAN_ROWS} строк по эвристике (макс. непустых строковых ячеек).")
    lines.append("- Имена колонок очищены: lower, trim, небуквенно-цифровые → `_`.")
    lines.append("- Файлы `~$*` (временные Excel) пропускались.")
    lines.append("- Ошибки отдельных файлов/листов пишутся в `errors.log`, обработка продолжается.")
    lines.append("")

    REPORT_PATH.write_text("\n".join(lines), encoding="utf-8")
    safe_print(f"Report written: {REPORT_PATH}")


# ===========================================================================
# Main
# ===========================================================================

def main() -> int:
    # Принудительно UTF-8 для stdout/stderr, если возможно (Python 3.7+)
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
    except Exception:
        pass

    safe_print("=" * 60)
    safe_print("agent_analyzer - start")
    safe_print(f"ROOT = {ROOT}")
    safe_print("=" * 60)

    # очистим errors.log для текущего прогона
    try:
        if ERRORS_LOG.exists():
            ERRORS_LOG.write_text("", encoding="utf-8")
    except OSError:
        pass

    # 2.1 Markdown
    safe_print("\n[1/4] Analyzing Markdown files...")
    try:
        insights, md_files = analyze_markdown_files()
        safe_print(f"  MD files: {len(md_files)}")
    except Exception as e:
        log_error(f"fatal md phase: {e}\n{traceback.format_exc()}")
        insights, md_files = (
            {"goals": [], "tasks": [], "processes": [], "problems": [], "solutions": [], "other": []},
            [],
        )

    # 2.2 Excel
    safe_print("\n[2/4] Analyzing Excel files...")
    try:
        sheets = analyze_all_excel()
        safe_print(f"  Sheets: {len(sheets)}")
    except Exception as e:
        log_error(f"fatal excel phase: {e}\n{traceback.format_exc()}")
        sheets = []

    # 2.3 Keys & relations
    safe_print("\n[3/4] Detecting keys and relationships...")
    try:
        natural_keys = detect_natural_keys(sheets)
        composite_key_md = propose_composite_key(natural_keys, sheets)
        relations = build_relationship_map(sheets)
        safe_print(f"  Natural key types: {len(natural_keys)}")
        safe_print(f"  Relations: {len(relations)}")
    except Exception as e:
        log_error(f"fatal keys/relations: {e}\n{traceback.format_exc()}")
        natural_keys = {}
        composite_key_md = "### Предложенный составной ключ\n\n_Ошибка при построении предложения._\n"
        relations = []

    # 2.4 Report
    safe_print("\n[4/4] Generating report...")
    try:
        generate_report(insights, md_files, sheets, natural_keys, composite_key_md, relations)
    except Exception as e:
        log_error(f"fatal report: {e}\n{traceback.format_exc()}")
        safe_print("FAILED to generate report", file=sys.stderr)
        return 1

    # stats
    err_count = 0
    if ERRORS_LOG.exists():
        content = ERRORS_LOG.read_text(encoding="utf-8").strip()
        if content:
            err_count = len(content.splitlines())

    safe_print("\n" + "=" * 60)
    safe_print("DONE")
    safe_print(f"  Excel files: {len({s['file'] for s in sheets})}")
    safe_print(f"  Sheets:      {len(sheets)}")
    safe_print(f"  MD files:    {len(md_files)}")
    safe_print(f"  Errors:      {err_count} (see errors.log)" if err_count else "  Errors:      0")
    safe_print(f"  Report:      {REPORT_PATH}")
    safe_print("=" * 60)
    return 0


if __name__ == "__main__":
    sys.exit(main())
