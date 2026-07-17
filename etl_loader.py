#!/usr/bin/env python
# -*- coding: utf-8 -*-
from __future__ import annotations

import logging
import math
import hashlib
import re
import sqlite3
import sys
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Callable, Iterable, Iterator, Sequence

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent
TARGET_DIR = PROJECT_ROOT / "upload"
# Аналитический warehouse в корне (как ожидает Next.js API)
DB_PATH = PROJECT_ROOT / "vsm_database.db"
LOG_PATH = PROJECT_ROOT / "db" / "etl_loader.log"

HEADER_SCAN_ROWS = 10
BATCH_SIZE = 5000
EXCEL_EPOCH = datetime(1899, 12, 30)
NA_VALUES = {"", "nan", "none", "null", "nat", "n/a", "n\\a", "-"}

LOGGER = logging.getLogger("etl_loader")


def setup_logging() -> logging.Logger:
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    logger = LOGGER
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    logger.propagate = False

    formatter = logging.Formatter(
        "%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    file_handler = logging.FileHandler(LOG_PATH, mode="w", encoding="utf-8")
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(formatter)

    stream_handler = logging.StreamHandler(sys.stdout)
    stream_handler.setLevel(logging.INFO)
    stream_handler.setFormatter(formatter)

    logger.addHandler(file_handler)
    logger.addHandler(stream_handler)
    return logger


def clean_column_name(name: Any) -> str:
    if name is None:
        return ""
    value = str(name).strip().lower()
    value = value.replace("ё", "е")
    value = re.sub(r"[^\w]+", "_", value, flags=re.UNICODE)
    value = re.sub(r"_+", "_", value)
    return value.strip("_")


def normalize_name(name: str | None) -> str:
    """Нормализация ФИО: верхний регистр, ё→е, схлопывание пробелов."""
    if not name:
        return ""
    text = str(name).upper().replace("Ё", "Е")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def generate_employee_uid(
    fio: str | None,
    birth_date: str | None,
    tab_number: str | None,
) -> str:
    """SHA256(ФИО|дата_рождения|табельный) — стабильный суррогатный ключ."""
    parts = [
        normalize_name(fio or ""),
        (str(birth_date).strip() if birth_date is not None else ""),
        (str(tab_number).strip() if tab_number is not None else ""),
    ]
    raw = "|".join(parts)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def extract_report_date_from_path(relative_path: str) -> str | None:
    """Достаёт дату учёта из имени файла: DD.MM.YYYY → YYYY-MM-DD."""
    name = Path(relative_path).name
    match = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", name)
    if not match:
        return None
    day, month, year = match.group(1), match.group(2), match.group(3)
    return f"{year}-{month}-{day}"


def normalize_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat(timespec="seconds")
    if isinstance(value, str):
        text = value.replace("\u00a0", " ").strip()
        return None if text.lower() in NA_VALUES else text
    return value


def transform_date(value: Any) -> str | None:
    value = normalize_value(value)
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        numeric = float(value)
        if math.isnan(numeric):
            return None
        if 0 < numeric < 60000:
            return (EXCEL_EPOCH + timedelta(days=numeric)).date().isoformat()
        text = str(int(numeric)) if numeric.is_integer() else str(numeric)
    else:
        text = str(value).strip()

    if not text or text.lower() in NA_VALUES:
        return None

    if re.fullmatch(r"\d+(?:[.,]\d+)?", text):
        numeric = float(text.replace(",", "."))
        if 0 < numeric < 60000:
            return (EXCEL_EPOCH + timedelta(days=numeric)).date().isoformat()

    normalized = text.replace("/", ".").replace("\\", ".").replace("T", " ")
    normalized = re.sub(r"\s+", " ", normalized).strip()

    for fmt in (
        "%d.%m.%Y",
        "%d.%m.%Y %H:%M",
        "%d.%m.%Y %H:%M:%S",
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y.%m.%d",
        "%Y.%m.%d %H:%M:%S",
        "%Y%m%d",
    ):
        try:
            parsed = datetime.strptime(normalized, fmt)
            return parsed.date().isoformat()
        except ValueError:
            continue

    try:
        parsed = datetime.fromisoformat(normalized)
        return parsed.date().isoformat()
    except ValueError:
        return text


def transform_money(value: Any) -> float | None:
    value = normalize_value(value)
    if value is None:
        return None

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        numeric = float(value)
        return None if math.isnan(numeric) else numeric

    text = str(value).strip().replace("\u00a0", " ").replace(" ", "")
    if not text or text.lower() in NA_VALUES:
        return None

    negative = text.startswith("(") and text.endswith(")")
    if negative:
        text = text[1:-1]

    text = text.replace("−", "-")
    text = re.sub(r"[^\d,.\-]", "", text)
    if not text:
        return None

    if text.count(",") and text.count("."):
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif text.count(",") > 1:
        text = text.replace(",", "")
    elif text.count(",") == 1:
        left, right = text.split(",", 1)
        text = f"{left}.{right}" if len(right) <= 3 else f"{left}{right}"
    elif text.count(".") > 1:
        parts = text.split(".")
        text = "".join(parts[:-1]) + "." + parts[-1]

    try:
        amount = float(text)
    except ValueError:
        return None
    return -amount if negative else amount


def _score_header_row(values: Sequence[Any]) -> float:
    if not values:
        return -1.0
    non_empty = 0
    textual = 0
    for value in values:
        normalized = normalize_value(value)
        if normalized is None:
            continue
        non_empty += 1
        if isinstance(normalized, str):
            try:
                float(normalized.replace(",", "."))
            except ValueError:
                textual += 1
    if non_empty == 0:
        return -1.0
    empty = len(values) - non_empty
    return (textual * 2.0) + non_empty - (empty * 0.25)


def find_header_row(rows: Sequence[Sequence[Any]]) -> int | None:
    best_row: int | None = None
    best_score = -1.0
    for row_number, row in enumerate(rows, start=1):
        score = _score_header_row(row)
        if score > best_score:
            best_score = score
            best_row = row_number
    return best_row if best_score > 0 else None


def _build_headers(row: Sequence[Any]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(row, start=1):
        column = clean_column_name(value) or f"unnamed_{index}"
        seen[column] = seen.get(column, 0) + 1
        headers.append(column if seen[column] == 1 else f"{column}_{seen[column]}")
    while headers and re.fullmatch(r"unnamed_\d+", headers[-1]):
        headers.pop()
    return headers


def find_files(base_dir: Path, patterns: Sequence[str]) -> list[Path]:
    compiled = [re.compile(pattern, re.IGNORECASE) for pattern in patterns]
    matches: list[Path] = []
    for path in base_dir.rglob("*"):
        if not path.is_file():
            continue
        if path.name.startswith("~$"):
            continue
        relative = path.relative_to(base_dir).as_posix()
        if any(regex.search(relative) for regex in compiled):
            matches.append(path)
    return sorted(matches, key=lambda item: item.as_posix().lower())


@dataclass
class WorkbookHandle:
    path: Path
    kind: str
    workbook: Any

    def sheet_names(self) -> list[str]:
        if self.kind == "xlsx":
            return list(self.workbook.sheetnames)
        return list(self.workbook.sheets)

    def iter_rows(self, sheet_name: str, max_rows: int | None = None) -> Iterator[list[Any]]:
        if self.kind == "xlsx":
            worksheet = self.workbook[sheet_name]
            for row in worksheet.iter_rows(max_row=max_rows, values_only=True):
                yield list(row)
            return

        with self.workbook.get_sheet(sheet_name) as sheet:
            for row_number, row in enumerate(sheet.rows(), start=1):
                if max_rows is not None and row_number > max_rows:
                    break
                yield [cell.v for cell in row]

    def close(self) -> None:
        try:
            self.workbook.close()
        except Exception:
            pass


def open_workbook(path: Path) -> WorkbookHandle:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return WorkbookHandle(
            path=path,
            kind="xlsx",
            workbook=load_workbook(path, read_only=True, data_only=True),
        )
    if suffix == ".xlsb":
        from pyxlsb import open_workbook as open_xlsb_workbook

        return WorkbookHandle(path=path, kind="xlsb", workbook=open_xlsb_workbook(str(path)))
    raise ValueError(f"Unsupported workbook format: {path}")


def resolve_sheets(
    workbook: WorkbookHandle,
    sheet_name: str | Sequence[str] | None = None,
    sheet_match: str | None = None,
) -> list[str]:
    available = workbook.sheet_names()
    lowered = {name.lower(): name for name in available}

    if isinstance(sheet_name, str):
        sheet_names = [sheet_name]
    else:
        sheet_names = list(sheet_name or [])

    resolved: list[str] = []
    for candidate in sheet_names:
        exact = lowered.get(candidate.lower())
        if exact:
            resolved.append(exact)

    if resolved:
        return resolved

    if sheet_match:
        regex = re.compile(sheet_match, re.IGNORECASE)
        resolved = [name for name in available if regex.search(name)]
        if resolved:
            return resolved

    if not sheet_names and not sheet_match:
        return available
    return []


def match_columns(headers: Sequence[str], mapping: dict[str, str]) -> tuple[list[tuple[int, str, str]], list[str]]:
    index_by_header = {clean_column_name(header): index for index, header in enumerate(headers)}
    matches: list[tuple[int, str, str]] = []
    missing: list[str] = []
    for source_column, target_column in mapping.items():
        cleaned_source = clean_column_name(source_column)
        if cleaned_source in index_by_header:
            matches.append((index_by_header[cleaned_source], cleaned_source, target_column))
        else:
            missing.append(cleaned_source)
    return matches, missing


def detect_event_type(path: Path) -> str:
    lowered = path.name.lower()
    if "прием" in lowered:
        return "hire"
    if "перевод" in lowered:
        return "transfer"
    if "уволь" in lowered:
        return "terminate"
    return "unknown"


def detect_direction(sheet_name: str) -> str:
    lowered = sheet_name.lower()
    if "прилет" in lowered or "прилёт" in lowered:
        return "arrival"
    if "вылет" in lowered:
        return "departure"
    return "unknown"


ETL_CONFIG: dict[str, dict[str, Any]] = {
    "dim_employee": {
        "patterns": [r"Базы[/\\]1с.*\.xlsx$", r"Маври[/\\]1с.*\.xlsx$"],
        "sheet_name": "Лист1",
        "key": "tab_number",
        "dedupe": True,
        "columns": {
            "табельный_номер_с_префиксами": "tab_number",
            "фио": "full_name",
            "дата_рождения": "birth_date",
            "удостоверение_серия": "passport_series",
            "удостоверение_номер": "passport_number",
            "организация": "organization",
            "подразделение": "department",
            "должность": "position",
            "разряд_категория": "grade",
            "состояние": "status",
            "график_работы": "work_schedule",
            "дата_приема": "hire_date",
            "дата_увольнения": "termination_date",
            "страна_гражданства": "citizenship",
            "территория": "territory",
            "место_рождения": "birth_place",
            "удостоверение_кем_выдан": "passport_issuer",
            "удостоверение_дата_выдачи": "passport_issue_date",
            "физическое_лицо_адрес_по_прописке": "address",
            "физическое_лицо_личный_мобильный_телефон": "mobile_phone",
        },
        "transformers": {
            "birth_date": transform_date,
            "hire_date": transform_date,
            "termination_date": transform_date,
            "passport_issue_date": transform_date,
        },
        "add_source_file": True,
        "add_employee_uid": True,
    },
    "dim_worksite": {
        "patterns": [r"Справочники[/\\]Площадки_Регион\.xlsx$"],
        "sheet_name": "Лист1",
        "key": "worksite_name",
        "dedupe": True,
        "columns": {
            "территория": "territory_code",
            "площадка": "worksite_name",
            "регион": "region",
        },
    },
    "dim_department": {
        "patterns": [r"Справочники[/\\]Подразделение_Участки\.xlsx$"],
        "sheet_name": ["Лист_1", "Лист1"],
        "key": "department_name",
        "dedupe": True,
        "columns": {
            "подразделение_база": "department_base",
            "подразделение": "department_name",
            "участок_ежедневный_учет": "daily_section",
        },
    },
    "dim_position": {
        "patterns": [
            r"Справочники[/\\]Классификация_по_должности\.xlsx$",
            r"Справочники[/\\]АУП_РОП_ИТР\.xlsx$",
        ],
        "sheet_name": ["Лист1", "Лист_1"],
        "dedupe": False,
        "columns": {
            "фактическая_должность": "position_name",
            "классификация": "classification",
            "должность": "position_name",
            "ауп_итр_роп": "aup_itr_rop",
        },
        "add_source_file": True,
    },
    "fact_daily_attendance": {
        "patterns": [
            r"Ежедневный[/\\](?!Новая папка).*Ежедневный учет.*\.xlsx$",
            r"Ежедневный[/\\]Новая папка[/\\]Ежедневный учет.*\.xlsx$",
        ],
        "sheet_name": "ЕЖЕДНЕВНЫЙ УЧЕТ",
        "dedupe": False,
        "columns": {
            "регион": "region",
            "площадка": "worksite_name",
            "табельный_номер": "tab_number",
            "фио": "full_name",
            "дата_рождения": "birth_date",
            "гражданство": "citizenship",
            "серия": "passport_series",
            "паспорт": "passport_number",
            "фактическая_должность": "position_name",
            "участок_отдел": "department",
            "виза": "visa",
            "вид_визы": "visa_type",
            "срок_до": "visa_until",
            "въезд_в_рф_прием_на_работу": "entry_or_hire",
        },
        "transformers": {
            "birth_date": transform_date,
            "visa_until": transform_date,
            "entry_or_hire": transform_date,
        },
        "add_source_file": True,
        "add_report_date": True,
        "add_employee_uid": True,
        "link_employee_uid": True,
    },
    "fact_hr_events": {
        "patterns": [r"Прием_Перевод\+Увольнеие[/\\](Прием|Переводы|Увольнение).*\.xlsx$"],
        "sheet_name": "Лист1",
        "dedupe": False,
        "columns": {
            "дата_события": "event_date",
            "компания": "company",
            "таб_номер": "tab_number",
            "сотрудник": "full_name",
            "организация_до": "prev_org",
            "территория_до": "prev_site",
            "подразделение_до": "prev_dept",
            "должность_до": "prev_pos",
            "разряд_до": "prev_grade",
            "организация_после": "new_org",
            "территория_после": "new_site",
            "подразделение_после": "new_dept",
            "должность_после": "new_pos",
            "разряд_после": "new_grade",
            "регистратор_события": "event_registrar",
        },
        "transformers": {"event_date": transform_date},
        "defaults": {"event_type": lambda ctx: detect_event_type(ctx["file_path"])},
        "add_source_file": True,
        "add_employee_uid": True,
        "link_employee_uid": True,
    },
    "fact_kpi": {
        "patterns": [r"Оценки[/\\].*Реестр.*\.xlsx$"],
        "sheet_match": r"реестр",
        "dedupe": False,
        "columns": {
            "таб": "tab_number",
            "фио": "full_name",
            "площадка": "worksite_name",
            "должность": "position_name",
            "подразделение_отдел": "department",
            "фио_мастера_того_кто_ставить_первую_оценку": "master_name",
            "средняя_оценка_1": "average_score_1",
            "фио_производителя_работ_того_кто_ставит_вторую_оценку": "foreman_name",
            "средняя_оценка_2": "average_score_2",
            "коэффициент_производственной_выроботка_берется_у_начальников_участка": "production_coefficient",
            "итоговая_оценка": "final_score",
        },
        "transformers": {
            "average_score_1": transform_money,
            "average_score_2": transform_money,
            "production_coefficient": transform_money,
            "final_score": transform_money,
        },
        "add_source_file": True,
        "add_employee_uid": True,
        "link_employee_uid": True,
    },
    "fact_ticket_finance": {
        "patterns": [r"Реестры по билетам[/\\].*\.xlsx$"],
        "sheet_name": ["Лист1", "Монтаж"],
        "dedupe": False,
        "columns": {
            "накладная": "invoice_number",
            "вид_услуги": "service_type",
            "подразделение": "department",
            "обоснование_перелета": "flight_reason",
            "организация": "organization",
            "классификация_сотрудников": "employee_classification",
            "операция": "operation",
            "ф_и_о": "full_name",
            "табельный_номер": "tab_number",
            "паспорт": "passport_number",
            "направление_город": "city",
            "маршрут": "route",
            "планируемая_дата_вылета": "planned_departure_date",
            "планируемая_дата_прилета": "planned_arrival_date",
            "дата_выписки_билета": "ticket_issue_date",
            "номер_билета": "ticket_number",
            "авиаперевозчик": "carrier",
            "поставщик": "supplier",
            "плательщик": "payer",
            "отдел": "cost_center",
            "сумма_билета_аг_вознаг": "ticket_amount_agency_fee",
            "стоимость_билета_стоимость_без_сбора": "base_ticket_cost",
            "сумма_ндс_билета": "vat_amount",
        },
        "transformers": {
            "planned_departure_date": transform_date,
            "planned_arrival_date": transform_date,
            "ticket_issue_date": transform_date,
            "ticket_amount_agency_fee": transform_money,
            "base_ticket_cost": transform_money,
            "vat_amount": transform_money,
        },
        "add_source_file": True,
    },
    "fact_flights": {
        "patterns": [r"Отчет_Прилет_Вылет[/\\].*\.(xlsx|xlsb)$"],
        "sheet_match": r"прилет|вылет|прилёт",
        "dedupe": False,
        "columns": {
            "таб": "tab_number",
            "проект": "project",
            "организация": "organization",
            "фио": "full_name",
            "фио_латиница": "full_name_latin",
            "дата_рождения": "birth_date",
            "гражданство": "citizenship",
            "серия_паспорта": "passport_series",
            "номер_паспорта": "passport_number",
            "рабочий_или_итр": "worker_category",
            "фактическая_должность": "position_name",
            "отдел_участок": "department",
            "отдел": "department",
            "дата_вылета_по_билету": "ticket_departure_date",
            "дата_прибытия": "arrival_date",
            "время_прибытия": "arrival_time",
            "дата_вылета_с_оп_по_заявке": "requested_departure_date",
            "дата_прибытия_в_пункт_назначения": "destination_arrival_date",
            "время_вылета": "departure_time",
            "дата_возвращения_на_оп": "return_to_worksite_date",
            "авиа_жд": "transport_type",
            "билет_куплен_да_нет_за_свой_счет": "ticket_flag",
            "обоснование_перелета": "travel_reason",
            "основание": "travel_reason",
            "сотрудник_прибыл_на_оп_не_прибыл_на_оп": "arrival_status",
            "сотрудник_убыл_с_оп_не_убыл_с_оп": "departure_status",
            "номер_телефона": "phone",
            "маршрут": "route",
            "примечание": "note",
            "вид_визы": "visa_type",
            "срок_действия_визы": "visa_valid_until",
            "номер_рейса": "flight_number",
            "сумма_стоимости_билета": "ticket_cost",
            "чартерный_рейс": "charter_flag",
            "чартерный_рейс_заявлен_да_нет": "charter_flag",
            "заявлен_на_чартер": "charter_requested",
            "дата_прибытия_в_новый_уренгой": "new_urengoy_arrival_date",
            "время_прибытия_в_новый_уренгой": "new_urengoy_arrival_time",
        },
        "transformers": {
            "birth_date": transform_date,
            "ticket_departure_date": transform_date,
            "arrival_date": transform_date,
            "requested_departure_date": transform_date,
            "destination_arrival_date": transform_date,
            "return_to_worksite_date": transform_date,
            "visa_valid_until": transform_date,
            "new_urengoy_arrival_date": transform_date,
            "ticket_cost": transform_money,
        },
        "defaults": {"direction": lambda ctx: detect_direction(ctx["sheet_name"])},
        "add_source_sheet": True,
    },
    "fact_employment": {
        "patterns": [
            r"Мара[/\\]ТРУДНОУСТРОЙСТВО 2025[/\\]2025 Трудоустройство таблица январь - июнь\.xlsx$"
        ],
        "sheet_name": "Трудоустройство",
        "dedupe": False,
        "columns": {
            "проект": "project",
            "таб": "tab_number",
            "фио": "full_name",
            "гражданство": "citizenship",
            "дата_рождения": "birth_date",
            "серия_паспорта": "passport_series",
            "номер_паспорта": "passport_number",
            "ропилиитр": "rop_or_itr",
            "фактическаядолжность": "position_name",
            "отдел_участок": "department",
            "планируемаядатаприбытиянаоп": "planned_arrival_at_site",
            "статус": "status",
            "примечание": "note",
        },
        "transformers": {
            "birth_date": transform_date,
            "planned_arrival_at_site": transform_date,
        },
    },
    "fact_visa": {
        "patterns": [r"Маври[/\\]Список граждан ДЗ с визами\.xlsx$"],
        "sheet_name": "Основа данных",
        "dedupe": False,
        "columns": {
            "фамилия_имя": "full_name",
            "паспорта": "passport_number",
            "гражданство": "citizenship",
            "дата_рождения": "birth_date",
            "срок_действия_паспорта": "passport_valid_until",
            "возраст": "age",
            "фактическая_должность": "position_name",
            "вид_работы": "work_type",
            "срок_действия_рнр": "rnr_valid_until",
            "должность_по_рнр": "rnr_position",
            "компания": "company",
            "табельный_номер_с_префиксами": "tab_number",
        },
        "transformers": {
            "birth_date": transform_date,
            "passport_valid_until": transform_date,
            "rnr_valid_until": transform_date,
        },
        "field_types": {"age": "REAL"},
    },
}

LOAD_ORDER = [
    "dim_worksite",
    "dim_department",
    "dim_position",
    "dim_employee",
    "fact_daily_attendance",
    "fact_hr_events",
    "fact_kpi",
    "fact_ticket_finance",
    "fact_flights",
    "fact_employment",
    "fact_visa",
]


def _ordered_unique(values: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for value in values:
        if value not in seen:
            seen.add(value)
            ordered.append(value)
    return ordered


def _table_columns(config: dict[str, Any]) -> list[str]:
    columns = _ordered_unique(config["columns"].values())
    columns.extend(col for col in config.get("defaults", {}) if col not in columns)
    if config.get("add_source_file") and "source_file" not in columns:
        columns.append("source_file")
    if config.get("add_source_sheet") and "source_sheet" not in columns:
        columns.append("source_sheet")
    if config.get("add_report_date") and "report_date" not in columns:
        columns.append("report_date")
    if config.get("add_employee_uid") and "employee_uid" not in columns:
        columns.append("employee_uid")
    return columns


def _create_indexes(conn: sqlite3.Connection, table_name: str, config: dict[str, Any]) -> None:
    if config.get("key"):
        key = config["key"]
        conn.execute(
            f'CREATE INDEX IF NOT EXISTS "ix_{table_name}_{key}" '
            f'ON {_quote(table_name)} ({_quote(key)})'
        )
    if config.get("add_employee_uid"):
        # уникальный индекс только для измерения сотрудников
        if table_name == "dim_employee":
            conn.execute(
                f'CREATE UNIQUE INDEX IF NOT EXISTS "ux_{table_name}_employee_uid" '
                f'ON {_quote(table_name)} ("employee_uid")'
            )
        else:
            conn.execute(
                f'CREATE INDEX IF NOT EXISTS "ix_{table_name}_employee_uid" '
                f'ON {_quote(table_name)} ("employee_uid")'
            )
    if config.get("add_report_date"):
        conn.execute(
            f'CREATE INDEX IF NOT EXISTS "ix_{table_name}_report_date" '
            f'ON {_quote(table_name)} ("report_date")'
        )
    conn.commit()


def resolve_fact_employee_uid(
    record: dict[str, Any],
    tab_to_uid: dict[str, str],
    uid_set: set[str],
) -> str | None:
    """
    Связь факта с dim_employee:
    1) по табельному номеру (надёжнее при расхождении ФИО),
    2) иначе по вычисленному SHA256 — только если uid уже есть в измерении.
    """
    tab = record.get("tab_number")
    tab_s = str(tab).strip() if tab is not None else ""
    if tab_s and tab_s in tab_to_uid:
        return tab_to_uid[tab_s]

    uid = generate_employee_uid(
        record.get("full_name"),
        record.get("birth_date"),
        tab_s,
    )
    if uid in uid_set:
        return uid
    return None


def _column_type(column_name: str, config: dict[str, Any]) -> str:
    explicit = config.get("field_types", {})
    if column_name in explicit:
        return explicit[column_name]
    transformer = config.get("transformers", {}).get(column_name)
    if transformer is transform_money:
        return "REAL"
    return "TEXT"


def _quote(identifier: str) -> str:
    return '"' + identifier.replace('"', '""') + '"'


def _create_table(conn: sqlite3.Connection, table_name: str, config: dict[str, Any]) -> list[str]:
    columns = _table_columns(config)
    conn.execute(f'DROP TABLE IF EXISTS {_quote(table_name)}')
    ddl_columns = ", ".join(
        f"{_quote(column_name)} {_column_type(column_name, config)}" for column_name in columns
    )
    conn.execute(f'CREATE TABLE {_quote(table_name)} ({ddl_columns})')
    conn.commit()
    return columns


def _apply_transform(
    target_column: str,
    value: Any,
    transformers: dict[str, Callable[[Any], Any]],
) -> Any:
    if target_column in transformers:
        return transformers[target_column](value)
    return normalize_value(value)


def load_table_from_files(
    table_name: str,
    config: dict[str, Any],
    base_dir: Path,
    conn: sqlite3.Connection,
    tab_to_uid: dict[str, str] | None = None,
    uid_set: set[str] | None = None,
) -> int:
    files = find_files(base_dir, config["patterns"])
    table_columns = _create_table(conn, table_name, config)
    transformers = config.get("transformers", {})
    defaults = config.get("defaults", {})
    key_column = config.get("key")
    dedupe_enabled = bool(key_column) and config.get("dedupe", True)
    seen_keys: set[Any] = set()
    inserted_rows = 0
    orphan_uid_rows = 0
    tab_map = tab_to_uid or {}
    uids = uid_set or set()

    log = LOGGER
    log.info("Loading %s from %s file(s)", table_name, len(files))

    if not files:
        log.warning("No files matched for %s", table_name)
        return 0

    # При дедупе first-wins: обрабатываем файлы от новых к старым
    # (обратный алфавитный порядок: Маври после Базы → Маври первый).
    if dedupe_enabled:
        files = list(reversed(files))

    insert_sql = (
        f'INSERT INTO {_quote(table_name)} '
        f'({", ".join(_quote(column) for column in table_columns)}) '
        f'VALUES ({", ".join("?" for _ in table_columns)})'
    )

    for file_path in files:
        relative_path = file_path.relative_to(base_dir).as_posix()
        report_date = (
            extract_report_date_from_path(relative_path)
            if config.get("add_report_date")
            else None
        )
        preview_workbook: WorkbookHandle | None = None
        metadata_by_sheet: dict[str, dict[str, Any]] = {}
        try:
            preview_workbook = open_workbook(file_path)
            candidate_sheets = resolve_sheets(
                preview_workbook,
                sheet_name=config.get("sheet_name"),
                sheet_match=config.get("sheet_match"),
            )
            if not candidate_sheets:
                log.warning("%s: no matching sheets in %s", table_name, relative_path)
                continue

            for sheet_name in candidate_sheets:
                preview_rows = list(preview_workbook.iter_rows(sheet_name, max_rows=HEADER_SCAN_ROWS))
                header_row_number = find_header_row(preview_rows)
                if header_row_number is None:
                    log.warning("%s: header not found in %s [%s]", table_name, relative_path, sheet_name)
                    continue

                headers = _build_headers(preview_rows[header_row_number - 1])
                matches, missing = match_columns(headers, config["columns"])
                if not matches:
                    log.warning(
                        "%s: no mapped columns found in %s [%s]",
                        table_name,
                        relative_path,
                        sheet_name,
                    )
                    continue

                if missing:
                    log.info(
                        "%s: %s missing mapped columns in %s [%s]",
                        table_name,
                        len(missing),
                        relative_path,
                        sheet_name,
                    )

                metadata_by_sheet[sheet_name] = {
                    "header_row_number": header_row_number,
                    "matches": matches,
                }
        except Exception:
            log.exception("%s: failed preview for %s", table_name, relative_path)
            continue
        finally:
            if preview_workbook is not None:
                preview_workbook.close()

        if not metadata_by_sheet:
            continue

        data_workbook: WorkbookHandle | None = None
        try:
            data_workbook = open_workbook(file_path)
            for sheet_name, meta in metadata_by_sheet.items():
                batch: list[tuple[Any, ...]] = []
                for row_number, row in enumerate(data_workbook.iter_rows(sheet_name), start=1):
                    if row_number <= meta["header_row_number"]:
                        continue

                    record = {column: None for column in table_columns}
                    mapped_value_present = False
                    for source_index, _source_name, target_name in meta["matches"]:
                        raw_value = row[source_index] if source_index < len(row) else None
                        value = _apply_transform(target_name, raw_value, transformers)
                        record[target_name] = value
                        if value is not None:
                            mapped_value_present = True

                    if not mapped_value_present:
                        continue

                    context = {
                        "file_path": file_path,
                        "relative_path": relative_path,
                        "sheet_name": sheet_name,
                        "row_number": row_number,
                        "record": record,
                    }
                    for column_name, factory in defaults.items():
                        try:
                            record[column_name] = factory(context)
                        except Exception:
                            log.exception(
                                "%s: default factory failed for %s [%s] row %s",
                                table_name,
                                relative_path,
                                sheet_name,
                                row_number,
                            )
                            record[column_name] = None

                    if config.get("add_source_file"):
                        record["source_file"] = relative_path
                    if config.get("add_source_sheet"):
                        record["source_sheet"] = sheet_name
                    if config.get("add_report_date"):
                        record["report_date"] = report_date

                    if config.get("add_employee_uid"):
                        if config.get("link_employee_uid"):
                            uid = resolve_fact_employee_uid(record, tab_map, uids)
                            if uid is None:
                                orphan_uid_rows += 1
                            record["employee_uid"] = uid
                        else:
                            record["employee_uid"] = generate_employee_uid(
                                record.get("full_name"),
                                record.get("birth_date"),
                                record.get("tab_number"),
                            )

                    if dedupe_enabled:
                        dedupe_key = record.get(key_column)
                        if dedupe_key is None or (
                            isinstance(dedupe_key, str) and not dedupe_key.strip()
                        ):
                            continue
                        if dedupe_key in seen_keys:
                            continue
                        seen_keys.add(dedupe_key)

                    batch.append(tuple(record.get(column) for column in table_columns))
                    if len(batch) >= BATCH_SIZE:
                        conn.executemany(insert_sql, batch)
                        conn.commit()
                        inserted_rows += len(batch)
                        batch.clear()

                if batch:
                    conn.executemany(insert_sql, batch)
                    conn.commit()
                    inserted_rows += len(batch)
        except Exception:
            log.exception("%s: failed load for %s", table_name, relative_path)
        finally:
            if data_workbook is not None:
                data_workbook.close()

    _create_indexes(conn, table_name, config)
    if orphan_uid_rows:
        log.warning(
            "%s: %s row(s) without matching employee_uid (orphan / unmatched)",
            table_name,
            orphan_uid_rows,
        )
    log.info("Loaded %s: %s row(s)", table_name, inserted_rows)
    return inserted_rows


def build_employee_maps(conn: sqlite3.Connection) -> tuple[dict[str, str], set[str]]:
    """tab_number → employee_uid и множество всех uid из dim_employee."""
    tab_to_uid: dict[str, str] = {}
    uid_set: set[str] = set()
    rows = conn.execute(
        'SELECT "tab_number", "employee_uid" FROM "dim_employee" '
        'WHERE "employee_uid" IS NOT NULL'
    ).fetchall()
    for tab, uid in rows:
        if uid:
            uid_set.add(uid)
        if tab and uid:
            tab_to_uid[str(tab).strip()] = uid
    return tab_to_uid, uid_set


def main() -> int:
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
    except Exception:
        pass

    log = setup_logging()
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    TARGET_DIR.mkdir(parents=True, exist_ok=True)
    LOG_PATH.parent.mkdir(parents=True, exist_ok=True)

    log.info("PROJECT_ROOT=%s", PROJECT_ROOT)
    log.info("TARGET_DIR=%s", TARGET_DIR)
    log.info("DB_PATH=%s", DB_PATH)
    log.info("LOG_PATH=%s", LOG_PATH)

    summary: dict[str, int] = {}
    failures: list[str] = []
    tab_to_uid: dict[str, str] = {}
    uid_set: set[str] = set()

    conn = sqlite3.connect(DB_PATH)
    try:
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")

        for table_name in LOAD_ORDER:
            try:
                cfg = ETL_CONFIG[table_name]
                if cfg.get("link_employee_uid"):
                    load_table_from_files(
                        table_name, cfg, TARGET_DIR, conn, tab_to_uid, uid_set
                    )
                else:
                    load_table_from_files(table_name, cfg, TARGET_DIR, conn)

                count = conn.execute(
                    f'SELECT COUNT(*) FROM {_quote(table_name)}'
                ).fetchone()[0]
                summary[table_name] = int(count)

                if table_name == "dim_employee":
                    tab_to_uid, uid_set = build_employee_maps(conn)
                    log.info(
                        "Employee maps ready: %s tabs, %s uids",
                        len(tab_to_uid),
                        len(uid_set),
                    )
            except Exception:
                log.exception("Table load failed: %s", table_name)
                failures.append(table_name)
                summary[table_name] = -1

        # Очистка мусора и пересборка витрин (agg_*)
        try:
            from warehouse_marts import clean_garbage_tab_numbers, update_aggregates

            cleaned = clean_garbage_tab_numbers(conn)
            log.info("Garbage tab_number removed: %s", cleaned)
            agg = update_aggregates(conn)
            log.info("Aggregates rebuilt: %s", agg)
        except Exception:
            log.exception("Aggregate rebuild failed")
            failures.append("aggregates")
    finally:
        conn.close()

    # зеркало для совместимости со старым путём
    legacy = PROJECT_ROOT / "db" / "project_data.db"
    try:
        if DB_PATH.exists():
            import shutil

            shutil.copy2(DB_PATH, legacy)
            log.info("Copied warehouse → %s", legacy)
    except Exception:
        log.exception("Failed to copy DB mirror to db/project_data.db")

    print("\nSUMMARY")
    for table_name in LOAD_ORDER:
        value = summary.get(table_name, -1)
        suffix = "FAILED" if value < 0 else str(value)
        print(f"{table_name}: {suffix}")

    if failures:
        log.error("Completed with failures: %s", ", ".join(failures))
        return 1

    log.info("ETL completed successfully")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
