# -*- coding: utf-8 -*-
"""
Скрипт для комплексного анализа директории с проектом.
Оптимизирован для работы с Excel-файлами >100k строк:
- чтение только первых строк для поиска заголовка;
- чтение сэмпла для типов данных;
- получение количества строк без загрузки всех данных.
"""

import os
import re
import logging
from collections import defaultdict, Counter
from pathlib import Path

import pandas as pd
import openpyxl
import xlrd

# ==================== КОНФИГУРАЦИЯ ====================
# !!! ИЗМЕНИТЕ НА ПУТЬ К ВАШЕЙ ПАПКЕ С ПРОЕКТОМ !!!
TARGET_DIR = r"C:\path\to\your\project"

MAX_ROWS_TO_SCAN = 100           # сколько первых строк сканировать для поиска заголовка
SAMPLE_ROWS_FOR_TYPES = 1000     # сколько строк читать для определения типов
REPORT_FILENAME = "PROJECT_ANALYSIS_REPORT.md"
LOG_FILENAME = "analysis_errors.log"

logging.basicConfig(
    filename=LOG_FILENAME,
    level=logging.ERROR,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

# ==================== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ====================

def find_files(directory, extensions):
    """Рекурсивно найти все файлы с заданными расширениями."""
    found = []
    for root, _, files in os.walk(directory):
        for file in files:
            if any(file.lower().endswith(ext) for ext in extensions):
                found.append(os.path.join(root, file))
    return found


def clean_column_name(name):
    """Привести имя колонки к единому формату."""
    if not isinstance(name, str):
        name = str(name)
    name = re.sub(r'\s+', ' ', name).strip()
    name = name.lower()
    name = re.sub(r'[^\w\s]', '', name)
    return name


def find_header_row(sheet_data, max_rows=MAX_ROWS_TO_SCAN):
    """
    Эвристический поиск строки-заголовка.
    sheet_data - список списков (первые max_rows строк).
    Возвращает индекс строки (0-based).
    """
    best_row = 0
    best_score = -1

    for i, row in enumerate(sheet_data[:max_rows]):
        if not row:
            continue

        non_empty = sum(1 for cell in row if cell is not None and str(cell).strip() != '')
        str_count = sum(1 for cell in row if isinstance(cell, str))
        num_count = sum(1 for cell in row if isinstance(cell, (int, float)))

        score = non_empty * 2 + str_count * 3 - num_count * 1
        if non_empty > 0:
            score += (non_empty / len(row)) * 10

        if score > best_score:
            best_score = score
            best_row = i

    return best_row


def read_excel_first_rows(file_path, sheet_name=None, max_rows=MAX_ROWS_TO_SCAN):
    """Прочитать первые max_rows строк из Excel-файла (без загрузки всего файла)."""
    ext = os.path.splitext(file_path)[1].lower()
    rows_data = []

    try:
        if ext == '.xlsx':
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet = wb[sheet_name] if sheet_name else wb.active
            for i, row in enumerate(sheet.iter_rows(values_only=True, max_row=max_rows)):
                rows_data.append(list(row))
            wb.close()
        elif ext == '.xls':
            wb = xlrd.open_workbook(file_path, formatting_info=False)
            sheet = wb.sheet_by_name(sheet_name) if sheet_name else wb.sheet_by_index(0)
            nrows = min(sheet.nrows, max_rows)
            for i in range(nrows):
                rows_data.append(sheet.row_values(i))
        else:
            raise ValueError(f"Unsupported extension: {ext}")
    except Exception as e:
        logging.error(f"Ошибка чтения первых строк {file_path} лист {sheet_name}: {e}")
        return []

    return rows_data


def get_total_rows(file_path, sheet_name):
    """Получить общее количество строк в листе без загрузки всех данных."""
    ext = os.path.splitext(file_path)[1].lower()
    try:
        if ext == '.xlsx':
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheet = wb[sheet_name] if sheet_name else wb.active
            total = sheet.max_row
            wb.close()
            return total
        elif ext == '.xls':
            wb = xlrd.open_workbook(file_path, formatting_info=False)
            sheet = wb.sheet_by_name(sheet_name) if sheet_name else wb.sheet_by_index(0)
            return sheet.nrows
    except Exception as e:
        logging.error(f"Ошибка получения количества строк {file_path} лист {sheet_name}: {e}")
        return 0
    return 0


def get_sheet_info(file_path, sheet_name, header_row_idx):
    """
    Получить информацию о листе: общее количество строк, колонки, типы.
    Читается только сэмпл (SAMPLE_ROWS_FOR_TYPES строк) для определения типов.
    """
    try:
        df_sample = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            header=header_row_idx,
            nrows=SAMPLE_ROWS_FOR_TYPES,
            engine=None
        )
        total_rows = get_total_rows(file_path, sheet_name)
        total_columns = len(df_sample.columns)
        dtypes = df_sample.dtypes.apply(lambda x: str(x)).to_dict()
        columns = [clean_column_name(col) for col in df_sample.columns]
        return {
            'total_rows': total_rows - (header_row_idx + 1),
            'total_columns': total_columns,
            'columns': columns,
            'dtypes': dtypes,
            'header_row': header_row_idx
        }
    except Exception as e:
        logging.error(f"Ошибка получения информации о листе {sheet_name} в {file_path}: {e}")
        return None


def parse_md_files(md_files):
    """Извлечь идеи из .md-файлов и сгенерировать сводку."""
    all_text = ""
    for md in md_files:
        try:
            with open(md, 'r', encoding='utf-8') as f:
                all_text += f.read() + "\n\n"
        except Exception as e:
            logging.error(f"Ошибка чтения {md}: {e}")

    if not all_text.strip():
        return "Не найдено описаний в .md файлах."

    keywords = ['идея', 'подход', 'бизнес', 'логик', 'цель', 'задач', 'решение', 'архитектура']
    sentences = re.findall(r'[^.!?]*[.!?]', all_text)
    relevant = [s.strip() for s in sentences if any(kw in s.lower() for kw in keywords)]

    summary = "### Сводка идей из предыдущих проектов\n\n"
    if relevant:
        summary += "Ключевые идеи:\n"
        for i, s in enumerate(relevant[:10], 1):
            summary += f"{i}. {s}\n"
    else:
        summary += all_text[:1000] + "...\n\n"

    summary += "### Предлагаемые улучшения\n\n"
    summary += """
- **Унификация данных**: привести все источники к единому формату.
- **Автоматизация**: внедрить ETL-пайплайн для регулярного обновления.
- **Контроль качества**: проверки на дубли, пропуски и аномалии.
- **Масштабируемость**: чанковая загрузка для больших объёмов.
- **Безопасность**: маскировка чувствительных данных (паспорт, ИНН).
"""
    return summary


# ==================== ОСНОВНАЯ ЛОГИКА ====================

def analyze_project():
    print(f"Начинаем анализ директории: {TARGET_DIR}")

    # ---- 1. .md файлы ----
    md_files = find_files(TARGET_DIR, ['.md'])
    print(f"Найдено .md файлов: {len(md_files)}")
    md_summary = parse_md_files(md_files)

    # ---- 2. Excel-файлы ----
    excel_files = find_files(TARGET_DIR, ['.xlsx', '.xls'])
    print(f"Найдено Excel-файлов: {len(excel_files)}")

    all_sheets_info = {}
    all_columns = defaultdict(list)
    column_frequency = Counter()
    potential_keys = set()

    id_keywords = [
        'табельный номер', 'табель', 'id', 'идентификатор', 'код',
        'фио', 'ф.и.о.', 'фамилия', 'имя', 'отчество',
        'паспорт', 'серия', 'номер', 'инн', 'снилс',
        'дата рождения', 'др', 'рождения'
    ]

    for file_path in excel_files:
        file_name = os.path.basename(file_path)
        print(f"Обработка файла: {file_name}")

        try:
            ext = os.path.splitext(file_path)[1].lower()
            if ext == '.xlsx':
                wb = openpyxl.load_workbook(file_path, read_only=True)
                sheet_names = wb.sheetnames
                wb.close()
            elif ext == '.xls':
                wb = xlrd.open_workbook(file_path, formatting_info=False)
                sheet_names = wb.sheet_names()
            else:
                continue

            for sheet_name in sheet_names:
                print(f"  Лист: {sheet_name}")
                try:
                    rows = read_excel_first_rows(file_path, sheet_name, MAX_ROWS_TO_SCAN)
                    if not rows:
                        continue

                    header_row_idx = find_header_row(rows)
                    header_row = rows[header_row_idx] if header_row_idx < len(rows) else []
                    if not header_row or all(c is None or str(c).strip() == '' for c in header_row):
                        header_row_idx = 0

                    sheet_info = get_sheet_info(file_path, sheet_name, header_row_idx)
                    if sheet_info is None:
                        continue

                    key = (file_name, sheet_name)
                    all_sheets_info[key] = {
                        'file_path': file_path,
                        'sheet_name': sheet_name,
                        'header_row': header_row_idx,
                        'total_rows': sheet_info['total_rows'],
                        'total_columns': sheet_info['total_columns'],
                        'columns': sheet_info['columns'],
                        'dtypes': sheet_info['dtypes']
                    }
                    all_columns[key] = sheet_info['columns']

                    for col in sheet_info['columns']:
                        column_frequency[col] += 1
                        col_clean = col.lower()
                        if any(kw in col_clean for kw in id_keywords):
                            potential_keys.add(col_clean)

                except Exception as e:
                    logging.error(f"Ошибка обработки листа {sheet_name} в {file_name}: {e}")
                    continue

        except Exception as e:
            logging.error(f"Ошибка открытия файла {file_path}: {e}")
            continue

    # ---- 3. Анализ ID и связей ----
    natural_keys = [col for col in potential_keys if column_frequency[col] > 1]

    fio_cols = [c for c in potential_keys if 'фио' in c or 'ф.и.о.' in c or 'фамилия' in c]
    birth_cols = [c for c in potential_keys if 'дата рождения' in c or 'др' in c or 'рождения' in c]
    passport_cols = [c for c in potential_keys if 'паспорт' in c or 'серия' in c or 'номер' in c]

    if fio_cols and (birth_cols or passport_cols):
        fio = fio_cols[0]
        if birth_cols:
            custom_id_suggestion = f"Конкатенация `{fio} + {birth_cols[0]}` (или хеш)"
        elif passport_cols:
            custom_id_suggestion = f"Конкатенация `{fio} + последние 4 цифры из {passport_cols[0]}`"
        else:
            custom_id_suggestion = f"Использовать `{fio}` (при условии уникальности)"
    else:
        custom_id_suggestion = "Рекомендуется создать суррогатный ключ (например, UUID)."

    # Связи (пересечения колонок)
    connections = defaultdict(list)
    file_list = list(all_sheets_info.keys())
    for i, (f1, s1) in enumerate(file_list):
        cols1 = set(all_columns[(f1, s1)])
        for f2, s2 in file_list[i+1:]:
            cols2 = set(all_columns[(f2, s2)])
            common = cols1.intersection(cols2)
            if common:
                connections[(f1, s1)].append((f2, s2, list(common)))

    # ---- 4. Генерация отчёта ----
    report_lines = []
    report_lines.append("# Аналитический отчёт по проекту\n")
    report_lines.append(f"**Дата**: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    report_lines.append(f"**Директория**: `{TARGET_DIR}`\n\n")

    report_lines.append("## 1. Улучшенная идея проекта\n")
    report_lines.append(md_summary)
    report_lines.append("\n")

    report_lines.append("## 2. Словарь данных (Data Dictionary)\n")
    report_lines.append("| Файл | Лист | Строка заголовка | Всего строк | Всего колонок | Колонки | Типы (пример) |\n")
    report_lines.append("|------|------|------------------|-------------|---------------|---------|---------------|\n")
    for (file_name, sheet_name), info in all_sheets_info.items():
        cols = ", ".join(info['columns'][:5])
        if len(info['columns']) > 5:
            cols += "..."
        dtypes = ", ".join([f"{k}: {v}" for k, v in list(info['dtypes'].items())[:5]])
        if len(info['dtypes']) > 5:
            dtypes += "..."
        report_lines.append(
            f"| {file_name} | {sheet_name} | {info['header_row']+1} | {info['total_rows']} | {info['total_columns']} | {cols} | {dtypes} |\n"
        )
    report_lines.append("\n")

    report_lines.append("## 3. Уникальные идентификаторы\n")
    report_lines.append("### Потенциальные естественные ключи\n")
    if natural_keys:
        for key in natural_keys:
            report_lines.append(f"- `{key}`\n")
    else:
        report_lines.append("Не обнаружено.\n")
    report_lines.append("\n### Предлагаемый кастомный ID\n")
    report_lines.append(f"**Рекомендация**: {custom_id_suggestion}\n\n")

    report_lines.append("## 4. Карта связей между файлами\n")
    if connections:
        for (f1, s1), linked in connections.items():
            for f2, s2, common in linked:
                report_lines.append(f"- `{f1}` (лист `{s1}`) <--> `{f2}` (лист `{s2}`) по: {', '.join(common)}\n")
    else:
        report_lines.append("Пересечений не найдено.\n")

    report_lines.append("\n### Частота колонок (топ-10)\n")
    for col, freq in column_frequency.most_common(10):
        report_lines.append(f"- `{col}`: {freq} раз(а)\n")

    report_text = "".join(report_lines)

    report_path = os.path.join(TARGET_DIR, REPORT_FILENAME)
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(report_text)

    print(f"Отчёт сохранён: {report_path}")
    print(f"Лог ошибок: {os.path.join(TARGET_DIR, LOG_FILENAME)}")
    return report_path


if __name__ == "__main__":
    if not os.path.exists(TARGET_DIR):
        print(f"Ошибка: директория {TARGET_DIR} не существует. Измените TARGET_DIR.")
    else:
        analyze_project()